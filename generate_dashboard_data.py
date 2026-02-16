"""
Generate JSON data from Excel for web dashboard
Run this after each scrape to update dashboard data
"""

import json
import re
from pathlib import Path
from datetime import datetime
from collections import Counter
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

SOURCE_FILE = "hemnet_listings.xlsx"
OUTPUT_FILE = "dashboard_data.json"


def parse_price(price_str):
    """Extract numeric price from string"""
    if not price_str:
        return None
    nums = re.findall(r'[\d\s]+', str(price_str).replace('\xa0', ' '))
    if nums:
        try:
            return int(''.join(nums[0].split()))
        except:
            return None
    return None


def parse_area(area_str):
    """Extract numeric area from string"""
    if not area_str:
        return None
    nums = re.findall(r'[\d,\.]+', str(area_str))
    if nums:
        try:
            return float(nums[0].replace(',', '.'))
        except:
            return None
    return None


def load_excel_data(filepath):
    """Load all data from Excel"""
    logger.info(f"Loading data from {filepath}")
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    data = []

    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col, header in enumerate(headers, 1):
            row_data[header] = ws.cell(row, col).value or ""
        data.append(row_data)

    wb.close()
    logger.info(f"Loaded {len(data)} listings")
    return data


def generate_time_series_data(data):
    """Generate time series data grouped by scraping date"""
    from collections import defaultdict

    # Group data by date
    data_by_date = defaultdict(list)
    for d in data:
        scraped_at = d.get('scraped_at', '')
        if scraped_at:
            # Extract date part (YYYY-MM-DD)
            date_str = str(scraped_at).split('T')[0] if 'T' in str(scraped_at) else str(scraped_at)
            data_by_date[date_str].append(d)

    # Generate time series
    time_series = {
        'dates': [],
        'total_listings': [],
        'listing_types': {'Bostad': [], 'New Construction': []},
        'listing_labels': {'Premium': [], 'Plus': [], 'Max': [], 'Standard': []},
        'listing_labels_pct': {'Premium': [], 'Plus': [], 'Max': [], 'Standard': []}
    }

    for date in sorted(data_by_date.keys()):
        listings = data_by_date[date]
        total = len(listings)

        time_series['dates'].append(date)
        time_series['total_listings'].append(total)

        # Count listing types
        bostad_count = sum(1 for l in listings if l.get('listing_type') == 'Bostad')
        newcon_count = sum(1 for l in listings if 'Nybyggnadsprojekt' in str(l.get('listing_type', '')))

        time_series['listing_types']['Bostad'].append(bostad_count)
        time_series['listing_types']['New Construction'].append(newcon_count)

        # Count listing labels
        premium_count = sum(1 for l in listings if l.get('listing_label') == 'Premium')
        plus_count = sum(1 for l in listings if l.get('listing_label') == 'Plus')
        max_count = sum(1 for l in listings if l.get('listing_label') == 'Max')
        standard_count = sum(1 for l in listings if not l.get('listing_label') or l.get('listing_label') == '')

        time_series['listing_labels']['Premium'].append(premium_count)
        time_series['listing_labels']['Plus'].append(plus_count)
        time_series['listing_labels']['Max'].append(max_count)
        time_series['listing_labels']['Standard'].append(standard_count)

        # Calculate percentages
        time_series['listing_labels_pct']['Premium'].append(round(premium_count / total * 100, 1) if total > 0 else 0)
        time_series['listing_labels_pct']['Plus'].append(round(plus_count / total * 100, 1) if total > 0 else 0)
        time_series['listing_labels_pct']['Max'].append(round(max_count / total * 100, 1) if total > 0 else 0)
        time_series['listing_labels_pct']['Standard'].append(round(standard_count / total * 100, 1) if total > 0 else 0)

    return time_series


def generate_dashboard_data(data):
    """Generate all analytics data for dashboard"""

    # Parse numeric values
    prices = [parse_price(d['price']) for d in data if parse_price(d['price'])]
    areas = [parse_area(d['area']) for d in data if parse_area(d['area'])]

    price_per_m2_list = []
    for d in data:
        p, a = parse_price(d['price']), parse_area(d['area'])
        if p and a and a > 0:
            price_per_m2_list.append(p / a)

    # Key Metrics
    key_metrics = {
        "total_listings": len(data),
        "avg_price": round(sum(prices) / len(prices)) if prices else 0,
        "median_price": sorted(prices)[len(prices)//2] if prices else 0,
        "min_price": min(prices) if prices else 0,
        "max_price": max(prices) if prices else 0,
        "avg_area": round(sum(areas) / len(areas), 1) if areas else 0,
        "avg_price_per_m2": round(sum(price_per_m2_list) / len(price_per_m2_list)) if price_per_m2_list else 0,
        "listings_with_price": len(prices),
        "listings_with_area": len(areas),
    }

    # Property Type Analysis
    property_types = Counter(d['property_type'] for d in data if d['property_type'])
    property_type_data = []
    for pt, count in property_types.most_common():
        pt_listings = [d for d in data if d['property_type'] == pt]
        pt_prices = [parse_price(d['price']) for d in pt_listings if parse_price(d['price'])]
        pt_areas = [parse_area(d['area']) for d in pt_listings if parse_area(d['area'])]

        pt_price_m2 = []
        for d in pt_listings:
            p, a = parse_price(d['price']), parse_area(d['area'])
            if p and a and a > 0:
                pt_price_m2.append(p / a)

        property_type_data.append({
            "name": pt,
            "count": count,
            "percentage": round(count * 100 / len(data), 1),
            "avg_price": round(sum(pt_prices) / len(pt_prices)) if pt_prices else 0,
            "avg_area": round(sum(pt_areas) / len(pt_areas), 1) if pt_areas else 0,
            "avg_price_m2": round(sum(pt_price_m2) / len(pt_price_m2)) if pt_price_m2 else 0,
        })

    # Location Analysis (Top 15)
    locations = Counter(d['location'] for d in data if d['location'])
    location_data = []
    for loc, count in locations.most_common(15):
        loc_listings = [d for d in data if d['location'] == loc]
        loc_prices = [parse_price(d['price']) for d in loc_listings if parse_price(d['price'])]
        location_data.append({
            "name": loc,
            "count": count,
            "avg_price": round(sum(loc_prices) / len(loc_prices)) if loc_prices else 0,
        })

    # Broker Analysis (Top 15)
    brokers = Counter(d['broker_name'] for d in data if d['broker_name'])
    broker_data = []
    for broker, count in brokers.most_common(15):
        broker_data.append({
            "name": broker,
            "count": count,
            "market_share": round(count * 100 / len(data), 1),
        })

    # Listing Labels
    labels = Counter(d['listing_label'] if d['listing_label'] else "Standard" for d in data)
    label_data = [{"name": label, "count": count} for label, count in labels.most_common()]

    # Rooms Distribution
    rooms = Counter(d['rooms'] for d in data if d['rooms'])
    rooms_data = [{"name": str(r), "count": c} for r, c in sorted(rooms.items(), key=lambda x: str(x[0]))]

    # Price Distribution (Buckets)
    price_ranges = [
        (0, 1000000, "< 1M"),
        (1000000, 2000000, "1-2M"),
        (2000000, 3000000, "2-3M"),
        (3000000, 4000000, "3-4M"),
        (4000000, 5000000, "4-5M"),
        (5000000, 7500000, "5-7.5M"),
        (7500000, 10000000, "7.5-10M"),
        (10000000, float('inf'), "> 10M"),
    ]
    price_distribution = []
    for min_p, max_p, label in price_ranges:
        count = sum(1 for p in prices if min_p <= p < max_p)
        price_distribution.append({"range": label, "count": count})

    # Features Analysis
    all_tags = []
    for d in data:
        if d['tags']:
            all_tags.extend([t.strip() for t in str(d['tags']).split(';')])
    tag_counts = Counter(all_tags)
    features_data = [
        {"name": tag, "count": count, "percentage": round(count * 100 / len(data), 1)}
        for tag, count in tag_counts.most_common(12)
    ]

    # Listing Type
    listing_types = Counter(d['listing_type'] for d in data if d['listing_type'])
    listing_type_data = [{"name": lt, "count": c} for lt, c in listing_types.most_common()]

    # Insights Generation
    insights = generate_insights(data, key_metrics, property_type_data, location_data, broker_data, features_data)

    # Recent Listings (last 10)
    recent_listings = []
    for d in data[-10:]:
        recent_listings.append({
            "id": d['id'],
            "title": d['title'],
            "location": d['location'],
            "price": d['price'],
            "area": d['area'],
            "rooms": d['rooms'],
            "property_type": d['property_type'],
            "listing_label": d['listing_label'] or "Standard",
        })

    # Time Series Data
    time_series = generate_time_series_data(data)

    return {
        "generated_at": datetime.now().isoformat(),
        "key_metrics": key_metrics,
        "property_types": property_type_data,
        "locations": location_data,
        "brokers": broker_data,
        "listing_labels": label_data,
        "rooms_distribution": rooms_data,
        "price_distribution": price_distribution,
        "features": features_data,
        "listing_types": listing_type_data,
        "insights": insights,
        "recent_listings": recent_listings,
        "time_series": time_series,
    }


def generate_insights(data, metrics, property_types, locations, brokers, features):
    """Generate analytical insights"""
    insights = []

    # Market Size Insight
    insights.append({
        "category": "Market Overview",
        "title": "Total Market Size",
        "text": f"Currently tracking {metrics['total_listings']} active listings with a combined market value of approximately {metrics['avg_price'] * metrics['total_listings'] / 1e9:.1f} billion kr.",
        "icon": "chart-line"
    })

    # Dominant Property Type
    if property_types:
        top_pt = property_types[0]
        insights.append({
            "category": "Property Mix",
            "title": "Dominant Property Type",
            "text": f"{top_pt['name']} dominates the market with {top_pt['count']} listings ({top_pt['percentage']}%), averaging {top_pt['avg_price']:,} kr.",
            "icon": "home"
        })

    # Price Analysis
    if metrics['avg_price'] and metrics['median_price']:
        diff = ((metrics['avg_price'] - metrics['median_price']) / metrics['median_price']) * 100
        if diff > 10:
            insights.append({
                "category": "Price Analysis",
                "title": "Price Skew Detected",
                "text": f"Average price ({metrics['avg_price']:,} kr) is {diff:.0f}% higher than median ({metrics['median_price']:,} kr), indicating presence of high-value luxury properties pulling up the average.",
                "icon": "trending-up"
            })

    # Hottest Location
    if locations:
        top_loc = locations[0]
        insights.append({
            "category": "Location Trends",
            "title": "Hottest Area",
            "text": f"{top_loc['name']} leads with {top_loc['count']} listings at average price of {top_loc['avg_price']:,} kr.",
            "icon": "map-pin"
        })

    # Top Broker
    if brokers:
        top_broker = brokers[0]
        insights.append({
            "category": "Broker Analysis",
            "title": "Market Leader",
            "text": f"{top_broker['name']} holds the largest market share at {top_broker['market_share']}% with {top_broker['count']} active listings.",
            "icon": "briefcase"
        })

    # Most Desired Feature
    if features:
        top_feature = features[0]
        insights.append({
            "category": "Features",
            "title": "Most Desired Feature",
            "text": f"{top_feature['name']} is the most common feature, present in {top_feature['percentage']}% of all listings.",
            "icon": "star"
        })

    # Price per m2 insight
    if property_types and len(property_types) > 1:
        sorted_by_m2 = sorted([pt for pt in property_types if pt['avg_price_m2'] > 0],
                              key=lambda x: x['avg_price_m2'], reverse=True)
        if sorted_by_m2:
            most_expensive = sorted_by_m2[0]
            insights.append({
                "category": "Value Analysis",
                "title": "Highest Price per m²",
                "text": f"{most_expensive['name']} has the highest price density at {most_expensive['avg_price_m2']:,} kr/m², making it the most premium property category.",
                "icon": "dollar-sign"
            })

    # Premium Listings
    premium_count = sum(1 for d in data if d.get('listing_label') == 'Premium')
    if premium_count > 0:
        premium_pct = premium_count * 100 / len(data)
        insights.append({
            "category": "Listing Quality",
            "title": "Premium Presence",
            "text": f"{premium_count} listings ({premium_pct:.0f}%) are marked as Premium, indicating sellers investing in higher visibility.",
            "icon": "award"
        })

    return insights


def main():
    source_path = Path(SOURCE_FILE)

    if not source_path.exists():
        logger.error(f"Source file not found: {SOURCE_FILE}")
        return

    # Load data
    data = load_excel_data(source_path)

    if not data:
        logger.error("No data found")
        return

    # Generate dashboard data
    logger.info("Generating dashboard data...")
    dashboard_data = generate_dashboard_data(data)

    # Save to JSON
    output_path = Path(OUTPUT_FILE)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    logger.info(f"Dashboard data saved to: {output_path}")


if __name__ == "__main__":
    main()
