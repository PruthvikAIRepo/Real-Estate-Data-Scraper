import json
import math
import time
import random
import logging
from datetime import datetime
from pathlib import Path
import cloudscraper
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('scraper.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Configuration
BASE_URL = "https://www.hemnet.se"
START_URL = "https://www.hemnet.se/bostader"
GRAPHQL_URL = "https://www.hemnet.se/graphql"
MUNICIPALITIES_FILE = "municipalities.json"
REQUEST_TIMEOUT = 30
MIN_DELAY = 2
MAX_DELAY = 4
EXCEL_FILE = "hemnet_listings.xlsx"

# Export Fields
EXPORT_FIELDS = [
    'id',
    'url',
    'listing_type',
    'property_type',
    'title',
    'location',
    'price',
    'area',
    'rooms',
    'plot_size',
    'viewing_date',
    'description',
    'broker_name',
    'broker_logo',
    'tags',
    'listing_label',
    'image_1',
    'image_2',
    'image_3',
    'image_4',
    'image_5',
    'scraped_at'
]

# Initialize scraper with browser spoofing
scraper = cloudscraper.create_scraper(
    browser={
        "browser": "chrome",
        "platform": "windows",
        "mobile": False
    }
)


def safe_get_text(element):
    """Safely extract text from an element, returning empty string if None."""
    if element is None:
        return ""
    return element.get_text(strip=True)


def safe_get_attr(element, attr):
    """Safely extract attribute from an element, returning empty string if None."""
    if element is None:
        return ""
    return element.get(attr, "")


def load_existing_data(filepath):
    """
    Load existing Excel file and return existing data with ID set for O(1) lookup.

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (existing_data list, existing_ids set)
    """
    if not filepath.exists():
        logger.info("No existing file found, starting fresh")
        return [], set()

    try:
        logger.info(f"Loading existing data from {filepath}")
        wb = load_workbook(filepath)
        ws = wb.active

        existing_data = []
        existing_ids = set()

        # Skip header row (row 1), read from row 2
        for row in range(2, ws.max_row + 1):
            row_data = {}
            for col, field in enumerate(EXPORT_FIELDS, 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data[field] = cell_value if cell_value is not None else ""

            existing_data.append(row_data)

            # Add ID to set for O(1) lookup
            listing_id = str(row_data.get('id', ''))
            if listing_id:
                existing_ids.add(listing_id)

        logger.info(f"Loaded {len(existing_data)} existing listings ({len(existing_ids)} unique IDs)")
        wb.close()
        return existing_data, existing_ids

    except Exception as e:
        logger.error(f"Error loading existing file: {e}")
        logger.warning("Starting fresh due to load error")
        return [], set()


def get_live_count(location_id):
    """
    Fetch current listing count for a municipality via Hemnet GraphQL API.

    Args:
        location_id: Hemnet municipality ID string

    Returns:
        Integer total listing count, or None on failure
    """
    query = '{ searchForSaleListings(limit: 0, search: { locationIds: ["%s"] }) { total } }' % location_id
    try:
        response = scraper.post(
            GRAPHQL_URL,
            json={"query": query},
            timeout=REQUEST_TIMEOUT
        )
        data = response.json()
        return data["data"]["searchForSaleListings"]["total"]
    except Exception as e:
        logger.warning(f"GraphQL count failed for location {location_id}: {e}")
        return None


def extract_listings(html, existing_ids=None):
    """
    Parse HTML and extract all property listing data with deduplication.

    Args:
        html: Raw HTML string from the page
        existing_ids: Set of existing listing IDs for O(1) duplicate lookup

    Returns:
        Tuple of (results list, new_count, duplicates_count)
    """
    if existing_ids is None:
        existing_ids = set()

    soup = BeautifulSoup(html, "lxml")

    # Find all property cards using dynamic class selector
    cards = soup.select("span[class*='Container_cardWrapper']")
    logger.info(f"Found {len(cards)} listing cards on page")

    results = []
    new_count = 0
    duplicates_count = 0
    scraped_at = datetime.now().isoformat()

    for card in cards:
        try:
            # Get the main link element - match both /bostad and /nybyggnadsprojekt
            link_tag = card.select_one("a[href^='/bostad'], a[href^='/nybyggnadsprojekt']")
            if not link_tag:
                continue

            # Basic identifiers
            href = safe_get_attr(link_tag, "href")
            url = urljoin(BASE_URL, href)
            listing_id = safe_get_attr(link_tag, "id")

            # O(1) duplicate check using Set
            if listing_id and listing_id in existing_ids:
                logger.debug(f"Skipping duplicate ID: {listing_id}")
                duplicates_count += 1
                continue

            # Determine listing type based on URL
            if href.startswith('/bostad'):
                listing_type = "Bostad"
            elif href.startswith('/nybyggnadsprojekt'):
                listing_type = "New Construction (Nybyggnadsprojekt)"
            else:
                listing_type = "Other"

            # Property Type (from SVG title inside location icon)
            property_type_svg = card.select_one("svg title")
            property_type = safe_get_text(property_type_svg)

            # Title (street address)
            title_el = card.select_one("h2")
            title = safe_get_text(title_el)

            # Location
            location_el = card.select_one("[data-testid='location-parsed-text']")
            location = safe_get_text(location_el)

            # Price
            price_el = card.select_one("span[class*='askingPrice']")
            price = safe_get_text(price_el)

            # Primary attributes container
            primary_attrs = card.select(
                "div[class*='ForSaleAttributes_primaryAttributesContainer'] "
                "div[class*='ForSaleAttributes_attribute']"
            )

            # Area (usually 2nd attribute after price)
            area = ""
            if len(primary_attrs) > 1:
                area = safe_get_text(primary_attrs[1])

            # Rooms (usually 3rd attribute)
            rooms = ""
            if len(primary_attrs) > 2:
                rooms = safe_get_text(primary_attrs[2])

            # Secondary attributes (plot size for villas, floor for apartments)
            secondary_attrs = card.select(
                "div[class*='ForSaleAttributes_secondaryAttributesContainer'] "
                "div[class*='ForSaleAttributes_attribute'] span"
            )
            plot_size = ""
            if secondary_attrs:
                plot_size = safe_get_text(secondary_attrs[0])

            # Viewing/Open House Date (NOT listing date)
            viewing_date_el = card.select_one("div[class*='NestDisplayTag_nestDisplayTagTypeOnImage']")
            viewing_date = safe_get_text(viewing_date_el)
            if viewing_date:
                viewing_date = viewing_date.strip()

            # Description
            description_el = card.select_one("div[class*='Description_descriptionContainer'] p")
            description = safe_get_text(description_el)

            # Broker name
            broker_name_el = card.select_one(
                "div[class*='BrokerInformation'] span[class*='NestBody']"
            )
            broker_name = safe_get_text(broker_name_el)

            # Broker logo URL
            broker_logo_el = card.select_one("div[class*='BrokerAgencyLogo'] img")
            broker_logo = safe_get_attr(broker_logo_el, "src")

            # Tags/Features (Balcony, Patio, etc.)
            tag_elements = card.select("div[class*='NestDisplayTag_nestDisplayTagTypeDefault']")
            tags = [safe_get_text(tag) for tag in tag_elements if safe_get_text(tag)]
            tags_str = "; ".join(tags)

            # Listing Label (Premium, Pro, etc.) - extract actual text
            listing_label_el = card.select_one("div[class*='NestDisplayTag_nestDisplayTagTypeTinted']")
            listing_label = safe_get_text(listing_label_el)

            # Images (up to 5)
            image_elements = card.select("img[data-testid='listing-card-image']")
            images = [safe_get_attr(img, "src") for img in image_elements if safe_get_attr(img, "src")]

            # Pad images list to always have 5 entries
            while len(images) < 5:
                images.append("")
            images = images[:5]

            # Build the result dictionary
            listing_data = {
                'id': listing_id,
                'url': url,
                'listing_type': listing_type,
                'property_type': property_type,
                'title': title,
                'location': location,
                'price': price,
                'area': area,
                'rooms': rooms,
                'plot_size': plot_size,
                'viewing_date': viewing_date,
                'description': description,
                'broker_name': broker_name,
                'broker_logo': broker_logo,
                'tags': tags_str,
                'listing_label': listing_label,
                'image_1': images[0],
                'image_2': images[1],
                'image_3': images[2],
                'image_4': images[3],
                'image_5': images[4],
                'scraped_at': scraped_at
            }

            results.append(listing_data)
            new_count += 1

            # Add to existing_ids set for future duplicate checks within same run
            if listing_id:
                existing_ids.add(listing_id)

            logger.debug(f"Extracted listing: {listing_id} - {title} ({listing_type})")

        except Exception as e:
            logger.warning(f"Skipping card due to error: {e}")
            continue

    return results, new_count, duplicates_count


def save_to_excel(listings, filename=None):
    """
    Save listings to an Excel file (.xlsx).

    Args:
        listings: List of listing dictionaries (can be existing + new combined)
        filename: Optional custom filename. Defaults to EXCEL_FILE constant.

    Returns:
        Path to the saved Excel file
    """
    if not listings:
        logger.warning("No listings to save")
        return None

    if filename is None:
        filename = EXCEL_FILE

    filepath = Path(filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hemnet Listings"

        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write header row
        for col_num, field in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=1, column=col_num, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_num, listing in enumerate(listings, 2):
            for col_num, field in enumerate(EXPORT_FIELDS, 1):
                value = listing.get(field, "")
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col_num, field in enumerate(EXPORT_FIELDS, 1):
            column_letter = get_column_letter(col_num)

            if field in ['url', 'broker_logo', 'image_1', 'image_2', 'image_3', 'image_4', 'image_5']:
                ws.column_dimensions[column_letter].width = 50
            elif field == 'description':
                ws.column_dimensions[column_letter].width = 60
            elif field in ['id', 'rooms', 'listing_label', 'listing_type']:
                ws.column_dimensions[column_letter].width = 16
            elif field in ['price', 'area', 'plot_size']:
                ws.column_dimensions[column_letter].width = 18
            else:
                ws.column_dimensions[column_letter].width = 25

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(filepath)
        logger.info(f"Saved {len(listings)} listings to {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        raise


def fetch_page(url):
    """
    Fetch a single page with error handling.

    Args:
        url: URL to fetch

    Returns:
        HTML content as string, or None if request failed
    """
    try:
        logger.info(f"Fetching: {url}")
        response = scraper.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()

        html = response.text
        logger.info(f"Received {len(html)} characters")
        return html

    except Exception as e:
        logger.error(f"Request failed: {e}")
        return None


def scrape_hemnet(output_file=None):
    """
    Main scraping function. Loops through all 290 Swedish municipalities,
    fetches live listing counts, and scrapes every page per municipality.

    Args:
        output_file: Optional custom output filename

    Returns:
        Tuple of (list of all new listings, path to Excel file)
    """
    filepath = Path(output_file) if output_file else Path(EXCEL_FILE)

    # Load municipalities from JSON (static list, built once)
    municipalities_path = Path(MUNICIPALITIES_FILE)
    if not municipalities_path.exists():
        logger.error(f"{MUNICIPALITIES_FILE} not found. Cannot run full scrape.")
        raise FileNotFoundError(f"{MUNICIPALITIES_FILE} not found")

    municipalities = json.loads(municipalities_path.read_text(encoding='utf-8'))
    total_municipalities = len(municipalities)

    logger.info(f"Loaded {total_municipalities} municipalities from {MUNICIPALITIES_FILE}")
    print(f"Loaded {total_municipalities} municipalities. Starting full scrape...", flush=True)

    # Load existing data with O(1) ID lookup set
    existing_data, existing_ids = load_existing_data(filepath)

    all_new_results = []
    total_new = 0
    total_duplicates = 0
    global_page = 0  # tracks overall page number for Flask progress

    for m_idx, municipality in enumerate(municipalities):
        loc_id = municipality["id"]
        loc_name = municipality["name"]

        # Fetch live listing count for this municipality
        total = get_live_count(loc_id)
        if total is None:
            logger.warning(f"Skipping {loc_name} — could not fetch count")
            print(f"Municipality {m_idx + 1}/{total_municipalities}: {loc_name} — skipped (count fetch failed)", flush=True)
            continue

        if total == 0:
            logger.info(f"Skipping {loc_name} — 0 listings")
            continue

        pages = math.ceil(total / 50)
        logger.info(f"[{m_idx + 1}/{total_municipalities}] {loc_name}: {total} listings across {pages} pages")
        print(f"Municipality {m_idx + 1}/{total_municipalities}: {loc_name} ({total} listings, {pages} pages)", flush=True)

        for page in range(1, pages + 1):
            global_page += 1
            url = f"{START_URL}?location_ids[]={loc_id}&page={page}"

            html = fetch_page(url)
            if html is None:
                logger.error(f"Failed to fetch {loc_name} page {page}, skipping")
                break

            listings, new_count, dup_count = extract_listings(html, existing_ids)

            if new_count == 0 and dup_count == 0:
                logger.info(f"No listings found on {loc_name} page {page}, stopping pagination")
                break

            all_new_results.extend(listings)
            total_new += new_count
            total_duplicates += dup_count

            logger.info(f"  Page {page}/{pages}: New: {new_count}, Duplicates: {dup_count}")
            # Keep "Page X:" format so Flask dashboard progress bar still works
            print(f"Page {global_page}: New: {new_count}, Duplicates: {dup_count}", flush=True)

            if page < pages:
                sleep_time = random.uniform(MIN_DELAY, MAX_DELAY)
                logger.info(f"Sleeping {sleep_time:.1f}s")
                time.sleep(sleep_time)

        # Checkpoint: save after every municipality so progress isn't lost on crash
        if all_new_results:
            all_listings = existing_data + all_new_results
            save_to_excel(all_listings, str(filepath))
            existing_data = existing_data + all_new_results
            all_new_results = []

            # Regenerate dashboard JSON so live count updates in UI
            try:
                from generate_dashboard_data import main as generate_dashboard
                generate_dashboard()
            except Exception as e:
                logger.warning(f"Dashboard JSON update failed: {e}")

        # Polite delay between municipalities
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    # Final summary
    output_path = filepath if filepath.exists() else None

    logger.info("=" * 50)
    logger.info("Full scrape complete!")
    logger.info(f"Total new listings added: {total_new}")
    logger.info(f"Total duplicates skipped: {total_duplicates}")
    logger.info(f"Total listings in file: {len(existing_data)}")
    if output_path:
        logger.info(f"Data saved to: {output_path}")

    print("=" * 50, flush=True)
    print("Scraping complete!", flush=True)
    print(f"New listings added: {total_new}", flush=True)
    print(f"Total listings in file: {len(existing_data)}", flush=True)

    return all_new_results, output_path


def main():
    """Entry point for the scraper."""
    try:
        new_listings, output_path = scrape_hemnet()

        if new_listings:
            logger.info("\nSample new record:")
            sample = new_listings[0]
            for key, value in sample.items():
                display_value = str(value)[:80] + "..." if len(str(value)) > 80 else value
                logger.info(f"  {key}: {display_value}")
        else:
            logger.info("No new listings added this run")

        # Auto-generate dashboard data
        try:
            from generate_dashboard_data import main as generate_dashboard
            logger.info("Generating dashboard data...")
            generate_dashboard()
            logger.info("Dashboard data updated! Open dashboard.html in browser.")
        except ImportError:
            logger.warning("Dashboard generator not found. Run generate_dashboard_data.py manually.")
        except Exception as e:
            logger.warning(f"Failed to generate dashboard: {e}")

    except KeyboardInterrupt:
        logger.info("\nScraping interrupted by user")
    except Exception as e:
        logger.error(f"Scraping failed: {e}")
        raise


if __name__ == "__main__":
    main()
